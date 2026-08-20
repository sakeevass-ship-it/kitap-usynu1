import os
from datetime import datetime
import pandas as pd
import streamlit as st
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
from openpyxl import Workbook, load_workbook

st.set_page_config(page_title='Кітап ұсыну жүйесі', page_icon='📚', layout='wide')
BASE = os.path.dirname(os.path.abspath(__file__))
BOOKS = os.path.join(BASE, 'books.csv')
SURVEY = os.path.join(BASE, 'survey_results.xlsx')

@st.cache_data
def load_books():
    return pd.read_csv(BOOKS)

def recommend(df, title, n=5):
    work = df.copy()
    work['features'] = (work['author'].fillna('') + ' ' + work['genre'].fillna('') + ' ' + work['description'].fillna(''))
    matrix = TfidfVectorizer().fit_transform(work['features'])
    sim = cosine_similarity(matrix)
    idxs = work.index[work['title'].str.lower() == title.lower()].tolist()
    if not idxs:
        return pd.DataFrame()
    idx = idxs[0]
    scores = sorted(enumerate(sim[idx]), key=lambda x: x[1], reverse=True)
    rows = []
    for i, score in scores:
        if i != idx:
            row = work.iloc[i].copy()
            row['similarity'] = score
            rows.append(row)
        if len(rows) >= n:
            break
    return pd.DataFrame(rows)

def ensure_survey():
    if not os.path.exists(SURVEY):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Жауаптар'
        ws.append(['Уақыты','Жас тобы','Оқу жиілігі','Формат','Жанрлар','Кітап таңдау тәсілі','Маңызды фактор','Кітап табу қиындығы','Ұсыну жүйесін қолдану','Ұсыныс негізі','Қазақша кітап функциясы','Қосымша ұсыныс'])
        wb.save(SURVEY)

def save_survey(values):
    ensure_survey()
    wb = load_workbook(SURVEY)
    ws = wb['Жауаптар']
    ws.append(values)
    wb.save(SURVEY)

books = load_books()
ensure_survey()

st.sidebar.title('📚 Мәзір')
page = st.sidebar.radio('Бөлімді таңдаңыз', ['🏠 Басты бет','📖 Кітаптар','✨ Кітап ұсыну','📝 Сауалнама','📊 Нәтижелер'])

if page == '🏠 Басты бет':
    st.title('📚 Кітап ұсыну жүйесі')
    st.subheader('Оқырман қызығушылығына сәйкес кітаптарды автоматты түрде ұсынатын қосымша')
    st.write('Бұл оқу жобасы кітаптардың авторы, жанры және сипаттамасын талдап, ұқсас кітаптарды TF-IDF және cosine similarity әдістерімен ұсынады.')
    c1,c2,c3 = st.columns(3)
    c1.metric('Кітап саны', len(books))
    c2.metric('Жанр саны', books['genre'].nunique())
    c3.metric('Орташа рейтинг', f"{books['rating'].mean():.1f} / 5")
    st.info('Сол жақ мәзірден кітаптарды қарап, ұсыныс алып немесе сауалнама толтыра аласыз.')

elif page == '📖 Кітаптар':
    st.title('📖 Кітаптар каталогы')
    query = st.text_input('🔎 Кітап немесе автор іздеу')
    genre = st.selectbox('Жанр', ['Барлығы'] + sorted(books['genre'].unique().tolist()))
    view = books.copy()
    if query:
        q=query.lower(); view=view[view['title'].str.lower().str.contains(q) | view['author'].str.lower().str.contains(q)]
    if genre != 'Барлығы': view=view[view['genre']==genre]
    for _, r in view.iterrows():
        with st.container(border=True):
            st.subheader(r['title'])
            st.write(f"**Автор:** {r['author']}  |  **Жанр:** {r['genre']}  |  ⭐ {r['rating']}")
            st.write(r['description'])

elif page == '✨ Кітап ұсыну':
    st.title('✨ Сізге кітап ұсынамыз')
    title = st.selectbox('Өзіңізге ұнаған кітапты таңдаңыз:', books['title'].tolist())
    count = st.slider('Ұсыныстар саны', 3, 8, 5)
    if st.button('Ұсыныс алу', type='primary'):
        recs = recommend(books, title, count)
        st.success(f'«{title}» кітабына ұқсас ұсыныстар:')
        for _, r in recs.iterrows():
            with st.container(border=True):
                st.subheader(r['title'])
                st.write(f"**{r['author']}** · {r['genre']} · ⭐ {r['rating']} · Ұқсастық: {r['similarity']*100:.0f}%")
                st.write(r['description'])

elif page == '📝 Сауалнама':
    st.title('📝 Кітап таңдау және кітап ұсыну жүйесін зерттеу')
    st.write('Мақсаты — оқырмандардың кітап таңдау әдеттері мен автоматты ұсыныс жүйесіне қажеттілігін анықтау. Жауаптар оқу-зерттеу мақсатында қолданылады.')
    with st.form('survey'):
        age=st.selectbox('1. Жас тобыңыз', ['15 жасқа дейін','15–18 жас','19–25 жас','26–35 жас','36–50 жас','50 жастан жоғары'])
        freq=st.selectbox('2. Кітапты қаншалықты жиі оқисыз?', ['Күн сайын','Аптасына бірнеше рет','Айына бірнеше рет','Сирек','Мүлдем оқымаймын'])
        fmt=st.selectbox('3. Қай форматты жиі қолданасыз?', ['Қағаз кітап','Электрондық кітап','Аудиокітап','Бірнеше формат'])
        genres=st.multiselect('4. Ұнататын жанрларыңыз', sorted(books['genre'].unique().tolist()) + ['Психология','Бизнес','Өмірбаян','Басқа'])
        choice=st.selectbox('5. Жаңа кітапты көбінесе қалай таңдайсыз?', ['Достар ұсынысы','Әлеуметтік желі','Рейтинг пен пікір','Авторына қарай','Жанрына қарай','Сипаттамасына қарай','Онлайн ұсыныс','Кездейсоқ'])
        factor=st.selectbox('6. Ең маңызды фактор', ['Жанры','Авторы','Сипаттамасы','Рейтингі','Оқырман пікірі','Танымалдылығы','Мұқабасы','Бағасы'])
        hard=st.selectbox('7. Қызықты кітапты табу қиынға соға ма?', ['Иә, жиі','Кейде','Сирек','Жоқ'])
        use=st.selectbox('8. Автоматты кітап ұсыну жүйесін қолданар ма едіңіз?', ['Иә','Мүмкін','Жоқ'])
        basis=st.multiselect('9. Жүйе қандай ақпарат негізінде ұсынсын?', ['Бұрын оқыған кітаптарым','Ұнататын жанрларым','Сүйікті авторларым','Бағаларым','Іздеу тарихым','Басқа пайдаланушылар бағалары'])
        kz=st.selectbox('10. Қазақ тіліндегі кітаптарды ұсыну функциясы қажет пе?', ['Иә','Жоқ','Маңызды емес'])
        extra=st.text_area('11. Қандай қосымша функция ұсынар едіңіз?')
        submitted=st.form_submit_button('Жауапты жіберу', type='primary')
    if submitted:
        save_survey([datetime.now().strftime('%Y-%m-%d %H:%M:%S'),age,freq,fmt,', '.join(genres),choice,factor,hard,use,', '.join(basis),kz,extra])
        st.success('Рақмет! Жауабыңыз Excel файлына сақталды.')

elif page == '📊 Нәтижелер':
    st.title('📊 Сауалнама нәтижелері')
    try:
        df = pd.read_excel(SURVEY)
    except Exception:
        df = pd.DataFrame()
    st.metric('Жауап саны', len(df))
    if len(df):
        st.subheader('Ұсыну жүйесін қолдану')
        st.bar_chart(df['Ұсыну жүйесін қолдану'].value_counts())
        st.subheader('Қазақша кітап функциясы')
        st.bar_chart(df['Қазақша кітап функциясы'].value_counts())
        st.dataframe(df, use_container_width=True)
    else:
        st.info('Әзірге сауалнама жауабы жоқ.')
    with open(SURVEY,'rb') as f:
        st.download_button('⬇️ Excel нәтижелерін жүктеу', f, file_name='survey_results.xlsx', mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

title,author,genre,description,rating
Harry Potter and the Philosopher's Stone,J. K. Rowling,Фэнтези,"Сиқыр мектебі, достық және шытырман оқиғалар туралы хикая.",4.8
The Hobbit,J. R. R. Tolkien,Фэнтези,"Қиял-ғажайып әлемдегі саяхат, айдаһар және батылдық туралы шығарма.",4.7
The Lord of the Rings,J. R. R. Tolkien,Фэнтези,"Сиқырлы сақина мен үлкен саяхат жайлы эпикалық фэнтези.",4.9
1984,George Orwell,Антиутопия,"Бақылау, билік және еркіндік мәселелерін көтеретін антиутопиялық роман.",4.7
Murder on the Orient Express,Agatha Christie,Детектив,"Пойыздағы жұмбақ кісі өлімін тергейтін әйгілі детективтік шығарма.",4.6
The Martian,Andy Weir,Ғылыми фантастика,"Марста жалғыз қалған астронавттың ғылым арқылы аман қалу күресі.",4.6
Dune,Frank Herbert,Ғылыми фантастика,"Шөлді ғаламшар, саясат және билік үшін күрес туралы эпикалық роман.",4.7
Pride and Prejudice,Jane Austen,Роман,"Махаббат, мінез және қоғам туралы классикалық роман.",4.6
The Alchemist,Paulo Coelho,Роман,"Арманын іздеген жас жігіттің рухани саяхаты туралы шығарма.",4.4
Sherlock Holmes: A Study in Scarlet,Arthur Conan Doyle,Детектив,"Шерлок Холмстың логика мен бақылауға негізделген алғашқы тергеулерінің бірі.",4.5
The Little Prince,Antoine de Saint-Exupery,Классика,"Достық, махаббат және адам болмысы туралы философиялық ертегі.",4.8
Abai Zholy,Mukhtar Auezov,Қазақ әдебиеті,"Абайдың өмірі мен қазақ қоғамын суреттейтін тарихи эпопея.",4.9
Koshpendiler,Ilyas Yesenberlin,Қазақ әдебиеті,"Қазақ хандығының тарихы мен дала өмірін бейнелейтін тарихи шығарма.",4.8
Менің атым Қожа,Бердібек Соқпақбаев,Қазақ әдебиеті,"Мектеп оқушысы Қожаның мінезі, арманы және қызықты оқиғалары туралы повесть.",4.8
Көшпенділер,Ілияс Есенберлин,Тарихи,"Қазақ хандығының қалыптасуы мен тарихи тұлғалар туралы трилогия.",4.9

streamlit>=1.30
pandas>=2.0
scikit-learn>=1.3
openpyxl>=3.1
